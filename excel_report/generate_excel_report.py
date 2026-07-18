import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.black_scholes import price_and_greeks
from app.core.iv_solver import solve_iv_chain
from app.core.interpretation import compute_pcr, pcr_card, compute_max_pain, max_pain_card, iv_spike_card
from app.core.pnl_decomposer import decompose_pnl
from app.data.mock_option_chain import generate_chain, generate_vol_surface
from app.data.mock_bnf_candles import generate_dataset
from app.core.backtester import run_backtest

NAVY = "1F2937"
ACCENT = "2563EB"
LIGHT = "EFF6FF"
FONT_NAME = "Arial"

header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
title_font = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
subtitle_font = Font(name=FONT_NAME, italic=True, size=10, color="6B7280")
body_font = Font(name=FONT_NAME, size=10)
thin_border = Border(bottom=Side(style="thin", color="D1D5DB"))


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=row[col])
    style_header_row(ws, start_row, len(df.columns))
    return start_row + len(df) + 1  # next free row


# ---------------------------------------------------------------------------
# 1. Compute everything from the engine
# ---------------------------------------------------------------------------
spot = 24350.0
chain = generate_chain(spot=spot, expiry_days=6)
T = chain.attrs["expiry_days"] / 365.0
r = 0.065

calls = chain[chain.option_type == "call"].reset_index(drop=True)
puts = chain[chain.option_type == "put"].reset_index(drop=True)

# Greeks for every strike (vectorized single call, not a loop)
call_pg = price_and_greeks(spot, calls.strike.values, T, r, calls.implied_volatility.values / 100, option_type="call")
put_pg = price_and_greeks(spot, puts.strike.values, T, r, puts.implied_volatility.values / 100, option_type="put")

calls["theoretical_price"] = call_pg["price"]
calls["delta"], calls["gamma"], calls["theta"], calls["vega"] = call_pg["delta"], call_pg["gamma"], call_pg["theta"], call_pg["vega"]
puts["theoretical_price"] = put_pg["price"]
puts["delta"], puts["gamma"], puts["theta"], puts["vega"] = put_pg["delta"], put_pg["gamma"], put_pg["theta"], put_pg["vega"]

# IV solver round-trip validation: solve IV back from theoretical price, should match input
solved_iv_calls = solve_iv_chain(calls.theoretical_price.values, spot, calls.strike.values, T, r, ["call"] * len(calls))
calls["iv_solver_check"] = np.round(solved_iv_calls * 100, 2)
calls["iv_solver_error"] = np.round(calls["iv_solver_check"] - calls["implied_volatility"], 4)

chain_full = pd.concat([calls, puts], ignore_index=True)
pcr = compute_pcr(chain_full)
max_pain_strike = compute_max_pain(chain_full)
pcr_note = pcr_card(pcr)
max_pain_note = max_pain_card(max_pain_strike, spot)
atm_iv = chain_full.iloc[(chain_full.strike - spot).abs().argsort()[:1]].implied_volatility.values[0]
iv_note = iv_spike_card(atm_iv, historical_avg_iv=13.5)

# Vol surface
surface = generate_vol_surface(spot=spot)
smile = surface[surface.expiry_days == surface.expiry_days.min()]
smile_calls = smile[smile.option_type == "call"][["strike", "implied_volatility"]].sort_values("strike")

# P&L decomposition: sample short strangle-style single position (short ATM call)
atm_strike = round(spot / 50) * 50
position = {"K": atm_strike, "r": r, "option_type": "call", "quantity": -50}  # short 1 lot (50)
snap_t0 = {"S": spot, "T": T, "sigma": 0.135}
snap_t1 = {"S": spot * 1.008, "T": T - 1 / 365, "sigma": 0.148}  # next day: spot +0.8%, IV +1.3pts
pnl = decompose_pnl(position, snap_t0, snap_t1)

# DOS backtest
bnf_data = generate_dataset(n_weeks=8)
trade_log, summary = run_backtest(bnf_data)

print("Computed all engine outputs. PCR:", round(pcr, 2), "Max pain:", max_pain_strike, "Backtest trades:", summary["total_trades"])

# ---------------------------------------------------------------------------
# 2. Build the workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# ---- Sheet 0: Overview ----
ws = wb.active
ws.title = "Overview"
ws["B2"] = "F&O Derivatives Analytics Platform"
ws["B2"].font = title_font
ws["B3"] = "AlgoLabs Assignment 2 — Engine Outputs & Backtest Results"
ws["B3"].font = subtitle_font

overview_rows = [
    ("Spot (NIFTY)", spot, ""),
    ("Nearest Expiry", f"{chain.attrs['expiry_days']} days", ""),
    ("Put-Call Ratio (PCR)", round(pcr, 2), pcr_note["note"]),
    ("Max Pain Strike", max_pain_strike, max_pain_note["note"]),
    ("ATM IV vs 30D Avg", f"{iv_note['value']}%", iv_note["note"]),
    ("DOS Backtest Win Rate", f"{summary['win_rate_pct']}%", f"{summary['total_trades']} trades over 8 weeks (Wed+Thu expiry sessions)"),
    ("DOS Backtest Total P&L", f"₹{summary['total_pnl_rupees']:,.0f}", f"Avg ₹{summary['avg_pnl_rupees']:,.0f}/trade, best ₹{summary['best_trade_rupees']:,.0f}, worst ₹{summary['worst_trade_rupees']:,.0f}"),
]
r0 = 5
ws.cell(row=r0, column=2, value="Metric").font = header_font
ws.cell(row=r0, column=3, value="Value").font = header_font
ws.cell(row=r0, column=4, value="Interpretation").font = header_font
for c in range(2, 5):
    ws.cell(row=r0, column=c).fill = header_fill
for i, (label, val, note) in enumerate(overview_rows):
    row = r0 + 1 + i
    ws.cell(row=row, column=2, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=row, column=3, value=val).font = body_font
    ws.cell(row=row, column=4, value=note).font = body_font
    ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r0 + 5].height = 30
ws.row_dimensions[r0 + 6].height = 30
autofit(ws, [3, 26, 16, 70])

note_row = r0 + len(overview_rows) + 3
ws.cell(row=note_row, column=2, value="Data note: option chain, vol surface, and Greeks use a realistic mock NSE-shaped chain "
                                       "(field names match the real NSE API). DOS backtest uses mock 5-min BNF futures candles "
                                       "since 5-min intraday history isn't available from NSE Bhav Copy (EOD-only) or the live "
                                       "chain endpoint. Swap the data/mock_*.py modules for live feeds without touching engine code.").font = subtitle_font
ws.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)
ws.row_dimensions[note_row].height = 45

# ---- Sheet 1: Option Chain & Greeks ----
ws1 = wb.create_sheet("Option Chain & Greeks")
chain_display = chain_full[["strike", "option_type", "last_price", "open_interest", "implied_volatility",
                             "theoretical_price", "delta", "gamma", "theta", "vega"]].sort_values(["strike", "option_type"])
chain_display.columns = ["Strike", "Type", "LTP", "Open Interest", "IV (%)", "BSM Price", "Delta", "Gamma", "Theta/day", "Vega (per 1% IV)"]
next_row = write_df(ws1, chain_display, start_row=2)
ws1["A1"] = f"NIFTY Option Chain — Spot {spot:,.0f}, Expiry {chain.attrs['expiry_days']}d"
ws1["A1"].font = title_font
autofit(ws1, [10, 8, 10, 14, 9, 11, 9, 9, 11, 14])
for row in ws1.iter_rows(min_row=3, max_row=next_row - 1):
    for cell in row:
        cell.font = body_font

# Chart: OI by strike, calls vs puts (visualizes PCR / positioning)
oi_pivot = chain_full.pivot_table(index="strike", columns="option_type", values="open_interest").reset_index()
oi_start = next_row + 2
write_df(ws1, oi_pivot.rename(columns={"strike": "Strike", "call": "Call OI", "put": "Put OI"}), start_row=oi_start)
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Open Interest by Strike (Call vs Put)"
chart1.y_axis.title = "Open Interest"
chart1.x_axis.title = "Strike"
n_oi = len(oi_pivot)
data = Reference(ws1, min_col=2, max_col=3, min_row=oi_start, max_row=oi_start + n_oi)
cats = Reference(ws1, min_col=1, min_row=oi_start + 1, max_row=oi_start + n_oi)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.width, chart1.height = 24, 11
ws1.add_chart(chart1, f"A{oi_start + n_oi + 3}")

# ---- Sheet 2: Volatility Smile & Surface ----
ws2 = wb.create_sheet("Vol Smile & Surface")
ws2["A1"] = "Implied Volatility Smile (Nearest Expiry) & Surface Data"
ws2["A1"].font = title_font
smile_disp = smile_calls.rename(columns={"strike": "Strike", "implied_volatility": "IV (%)"})
next_row2 = write_df(ws2, smile_disp, start_row=3)
chart2 = LineChart()
chart2.title = "IV Smile — Call IV vs Strike (nearest expiry)"
chart2.y_axis.title = "Implied Volatility (%)"
chart2.x_axis.title = "Strike"
data2 = Reference(ws2, min_col=2, min_row=3, max_row=next_row2 - 1)
cats2 = Reference(ws2, min_col=1, min_row=4, max_row=next_row2 - 1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width, chart2.height = 22, 10
ws2.add_chart(chart2, "D3")

surf_start = next_row2 + 2
surf_disp = surface[surface.option_type == "call"].pivot_table(index="strike", columns="expiry_days", values="implied_volatility")
surf_disp.columns = [f"{d}D expiry" for d in surf_disp.columns]
surf_disp = surf_disp.reset_index().rename(columns={"strike": "Strike"})
ws2.cell(row=surf_start - 1, column=1, value="Vol Surface Grid (Call IV %, Strike x Expiry)").font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
write_df(ws2, surf_disp, start_row=surf_start)
autofit(ws2, [10, 10, 10, 10, 10, 10, 10])

# ---- Sheet 3: Interpretation Cards ----
ws3 = wb.create_sheet("Interpretation Cards")
ws3["A1"] = "Plain-Language Interpretation Cards"
ws3["A1"].font = title_font
cards = [
    ("PCR Signal", pcr_note["sentiment"], f"{pcr_note['value']}", pcr_note["note"]),
    ("Max Pain", "—", f"{max_pain_note['value']:.0f}", max_pain_note["note"]),
    ("IV Spike", "—" if iv_note["value"] is None else ("Elevated" if iv_note["value"] > 15 else "Depressed" if iv_note["value"] < -15 else "Normal"),
     f"{iv_note['value']}%", iv_note["note"]),
]
headers = ["Card", "Signal", "Value", "Plain-Language Note"]
for j, h in enumerate(headers):
    ws3.cell(row=3, column=1 + j, value=h)
style_header_row(ws3, 3, 4)
for i, (card, signal, value, note) in enumerate(cards):
    row = 4 + i
    ws3.cell(row=row, column=1, value=card).font = Font(name=FONT_NAME, bold=True, size=10)
    ws3.cell(row=row, column=2, value=signal).font = body_font
    ws3.cell(row=row, column=3, value=value).font = body_font
    ws3.cell(row=row, column=4, value=note).font = body_font
    ws3.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[row].height = 45
autofit(ws3, [14, 12, 10, 90])

# ---- Sheet 4: P&L Decomposition ----
ws4 = wb.create_sheet("P&L Decomposer")
ws4["A1"] = "P&L Decomposition — Sample Position (Short 1 lot ATM NIFTY Call)"
ws4["A1"].font = title_font
ws4["A2"] = f"Position: Short {abs(position['quantity'])} qty of {position['K']} Call | Spot moves {spot:,.0f} -> {snap_t1['S']:,.0f} (+0.8%), IV {snap_t0['sigma']*100:.1f}% -> {snap_t1['sigma']*100:.1f}%"
ws4["A2"].font = subtitle_font

pnl_rows = pd.DataFrame([
    {"Component": "Delta P&L", "Contribution (₹)": round(pnl["delta_pnl"], 2)},
    {"Component": "Gamma P&L", "Contribution (₹)": round(pnl["gamma_pnl"], 2)},
    {"Component": "Theta P&L", "Contribution (₹)": round(pnl["theta_pnl"], 2)},
    {"Component": "Vega P&L", "Contribution (₹)": round(pnl["vega_pnl"], 2)},
    {"Component": "Residual (higher-order)", "Contribution (₹)": round(pnl["residual_pnl"], 2)},
    {"Component": "Actual Total P&L", "Contribution (₹)": round(pnl["actual_pnl"], 2)},
])
next_row4 = write_df(ws4, pnl_rows, start_row=4)
ws4.cell(row=next_row4 + 1, column=1, value=f"Primary driver of today's P&L: {pnl['primary_driver']}").font = Font(name=FONT_NAME, bold=True, size=11, color=ACCENT)
autofit(ws4, [26, 18])

chart4 = BarChart()
chart4.type = "col"
chart4.title = "P&L Attribution by Greek"
chart4.y_axis.title = "₹"
data4 = Reference(ws4, min_col=2, min_row=4, max_row=4 + 4)  # exclude "Actual Total" row from the attribution bars
cats4 = Reference(ws4, min_col=1, min_row=5, max_row=4 + 4)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.width, chart4.height = 20, 10
ws4.add_chart(chart4, "A12")

# ---- Sheet 5: DOS Backtest ----
ws5 = wb.create_sheet("DOS Backtest")
ws5["A1"] = "DOS (Direction of SuperTrend) Strategy — Backtest Results"
ws5["A1"].font = title_font
ws5["A2"] = "Bank Nifty Futures, 5-min SuperTrend(10,3), Wed/Thu expiry sessions, 8 weeks"
ws5["A2"].font = subtitle_font

summary_rows = pd.DataFrame([{"Metric": k.replace("_", " ").title(), "Value": v} for k, v in summary.items()])
next_row5 = write_df(ws5, summary_rows, start_row=4)
autofit(ws5, [22, 14])

log_start = next_row5 + 2
ws5.cell(row=log_start - 1, column=1, value="Trade Log").font = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
log_disp = trade_log[["session_date", "day_type", "entry_time", "exit_time", "option_type", "strike",
                       "premium_sold", "premium_exit", "exit_reason", "pnl_rupees", "cumulative_pnl"]].copy()
log_disp.columns = ["Date", "Day", "Entry Time", "Exit Time", "Type", "Strike", "Premium Sold", "Premium Exit", "Exit Reason", "P&L (₹)", "Cumulative P&L (₹)"]
log_disp["Date"] = log_disp["Date"].astype(str)
log_disp["Entry Time"] = log_disp["Entry Time"].astype(str)
log_disp["Exit Time"] = log_disp["Exit Time"].astype(str)
next_row_log = write_df(ws5, log_disp, start_row=log_start)
for row in ws5.iter_rows(min_row=log_start + 1, max_row=next_row_log - 1):
    for cell in row:
        cell.font = body_font
autofit_widths = [11, 10, 16, 16, 6, 8, 12, 12, 12, 11, 16]
for i, w in enumerate(autofit_widths, start=1):
    ws5.column_dimensions[get_column_letter(i)].width = w

chart5 = LineChart()
chart5.title = "Equity Curve — Cumulative P&L (₹)"
chart5.y_axis.title = "Cumulative P&L (₹)"
chart5.x_axis.title = "Trade #"
data5 = Reference(ws5, min_col=11, min_row=log_start, max_row=next_row_log - 1)
chart5.add_data(data5, titles_from_data=True)
chart5.width, chart5.height = 24, 11
ws5.add_chart(chart5, f"A{next_row_log + 2}")

out_path = "/mnt/user-data/outputs/FnO_Analytics_Report.xlsx"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"Saved workbook to {out_path}")
